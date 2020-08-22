# coding: utf-8

###############################################################################
######             TRABALHO PRÁTICO : ALGORITMO EM GRAFOS                ######
######                   TEMA: COLORAÇÃO EM GRAFOS                       ######
######                                                                   ######
######                      Python Version: 3.4                          ######
######                                                                   ######
######      KELLYSON SANTOS DA SILVA - 201820366 - 10A                   ######
######      LAYSE CRISTINA SILVA GARCIA - 201811177 - 10A                ######
###############################################################################
'''------------------------------------------------------------Imports------------------------------------------------------------'''
import openpyxl #install from: pip install openpyxl
import os.path
from Pessoa import Pessoa
from enum.Enums import Enums
'''-------------------------------------------------------------------------------------------------------------------------------'''
Font = openpyxl.styles.Font
#Border = openpyxl.styles.Border


'''
    A classe XlsReaderHelper é a classe que auxilia na leitura da planilha de
    respostas, já criando objetos da Classe Pessoa e retornando a lista de pessoas
'''
class XlsReaderHelper:

    # Construtor da classe inicializa o leitor do arquivo
    def __init__(self, arquivo, aba):
        workBook = openpyxl.load_workbook(arquivo)
        self.workSheet = workBook[aba]
        self.maxRow =  self.workSheet.max_row

    # Célula: Captura linha e coluna da planilha
    def getCell(self, linha, coluna):
        return self.workSheet.cell(row = linha, column = coluna).value

    # Percorre a planilha a partir da segunda linha e segunda coluna, devido as
    # informações coletadas no formulário. Faz o tratamento para ' - ' presentes
    # em algumas colunas, criando assim um dicionário
    def buscarPessoas(self):

        pessoas = []

        integer = 2
        for linha in range(2, self.maxRow + 1): # +1 não inclusivo

            # Pega as informações de cada coluna da planilha
            idade = self.getCell(linha, 2)
            sexo = self.getCell(linha, 3)
            cidade = self.getCell(linha, 4)
            quadroDeRisco = self.getCell(linha, 5)
            tempoMedio = self.getCell(linha, 6)

            horarioDePreferencia = str(self.getCell(linha, 7)).split(' - ')
            horarioDePreferencia = toInteger(horarioDePreferencia)

            diasDaSemana = str(self.getCell(linha, 8)).split(' - ')
            diasDaSemana = toInteger(diasDaSemana)

            sintomas = self.getCell(linha, 9)

            # Cria o objeto Pessoa
            pessoa = Pessoa(linha - 1, idade, sexo, cidade, quadroDeRisco, tempoMedio, horarioDePreferencia, diasDaSemana, sintomas)
            # Adiciona a pessoa criada na lista de pessoas e retorna a lista
            pessoas.append(pessoa)

        return pessoas

# Cria o dicionário
def toInteger(lista):
    return list(map(int, lista))

class XlsWriterHelper:
    # Construtor da classe inicializa o escritor do arquivo
    def __init__(self, nomePlanilhaResultado):
        self.nomePlanilhaResultado = nomePlanilhaResultado

    # Gera Planilha alocando os tipos de pessoa com os horários,
    # de acordo com o Algoritmo de Coloração
    def CriaPlanilha(self, schedule, nomeDaAba = 'Todos'):

        # Inicializa o arquivo como nenhum
        arquivo = None

        # Verifica se o arquivo já existe no disco
        arquivoExistente = os.path.isfile(nomePlanilhaResultado)

        if arquivoExistente:
            # Carrega o arquivo do disco
            arquivo = openpyxl.load_workbook(filename = nomePlanilhaResultado)
        else:
            # Cria um novo arquivo
            arquivo = openpyxl.Workbook()
            # Apaga a planilha já configurada ao criar o arquivo
            del arquivo["Sheet"]

        # Para cada vértice colorido, realiza os processos
        for vertice in self.Lista_Coloridos:

            # Verifica se a aba ainda não existe
            if not (str(nomeDaAba) in arquivo.sheetnames):

                # Cria a nova aba de acordo com o valor por parâmetro
                arquivo.create_sheet(str(nomeDaAba))
                planilha = arquivo[str(nomeDaAba)]

                # Adiciona os Dias da Semana na Planilha
                for i in range(2, 8):
                    planilha.cell(row = 1, column = i).value = ENUMS.DIAS[i - 2]
                    # Configura a célula de dias da semana
                    planilha.cell(row = 1, column = i).alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')

                # Adiciona os horários de funcionamento do mercado, de acordo com a pesquisa
                for i in range(2, len(Enums.HORARIOS) + 2):
                    planilha.cell(row = i, column = 1).value = Enums.HORARIOS[i]
                    # Configura a célula de horários de funcionamento
                    planilha.cell(row = i, column = 1).alignment = openpyxl.styles.Alignment(horizontal = 'right', vertical = 'center')

                for i in range(1, len(Enums.HORARIOS) + 2):
                    for j in range(1, 7):
                        # Redimensiona as linhas e colunas
                        planilha.row_dimensions[i].height = 40
                        planilha.column_dimensions[ENUMS.DIA_COLUNA[j - 2]].width = 60

                        # Torna negrito
                        planilha.row_dimensions[i].font = Font(bold = True)
                        planilha.column_dimensions[ENUMS.DIA_COLUNA[j - 2]].font = Font(bold = True)

            # Se a aba existir pula o processo acima
            planilha = arquivo[str(nomeDaAba)]

                ## cenas do proximo capitulo
            ######### DEPENDE DA ESTRUTURA UTILIZADA ############
            # Encontra as posições das células na planilha pelo dia da semana e o horário do vértice
            horario = self.obtemHorario(vertice.cor)
            j = diaDaSemana.index(horario[1]) + 2
            i = int((vertice.cor - (j - 2))/5) + 2
            # Altera o valor da célula, colocando a matéria e o professor no horário em questão
            ValorCell = 'Grupo de Risco: ' + str(vertice.materia) + ' | Sexo: ' + vertice.professor + ' | Idade: ' + vertice.professor + ' | Cidade:  ' + str(vertice.aula)
            planilha.cell(row = i, column = j).value = ValorCell
            # Formata a célula
            planilha.cell(row = i, column = j).alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')

        # Cria uma Lista de Planilhas(abas) do arquivo
        ListaDePlanilhas = arquivo.sheetnames
        # Para cada planilha existente, adiciona-se "Aba" antes do nome
        for Planilha in ListaDePlanilhas:
            planilha = arquivo[Planilha]
            NomeNovoDaPlanilha = 'Aba: ' + nomeDaAba
            planilha.title = NomeNovoDaPlanilha

        # Ordena as planilhas por ordem alfabetica antes de salvar
        # arquivo._sheets.sort(key=lambda ws: ws.title)

        # Salva o arquivo e o fecha
        arquivo.save(nomePlanilhaResultado)

        # Se o arquivo não for fechado, todo esse método não é executado
        arquivo.close()
