# coding: utf-8

###############################################################################
######             TRABALHO PRÁTICO : ALGORITMO EM GRAFOS                ######
######                   TEMA: COLORAÇÃO EM GRAFOS                       ######
######                                                                   ######
######                      Python Version: 2.7                          ######
######                                                                   ######
######      KELLYSON SANTOS DA SILVA - 201820366 - 10A                   ######
######      LAYSE CRISTINA SILVA GARCIA - 201811177 - 10A                ######
###############################################################################
'''------------------------------------------------------------Imports------------------------------------------------------------'''
import openpyxl #install from: pip install openpyxl
import os.path
from Pessoa import Pessoa
from enum.Enums import Enums
'''------------------------------------------------Objetos de Estilo da Planilha--------------------------------------------------'''
Font = openpyxl.styles.Font
PatternFill = openpyxl.styles.PatternFill
Alignment = openpyxl.styles.Alignment
Border = openpyxl.styles.Border
Side = openpyxl.styles.Side
'''-------------------------------------------------------------------------------------------------------------------------------'''

'''
    A classe XlsHelper é a classe que auxilia na leitura e escrita da planilha.
    Possui duas classes, como já dito, uma de leitura e outra de escrita.
        - Reader: faz a leitura do arquivo, criando objetos da Classe Pessoa e
          retornando a lista de pessoas;
        - Writer: faz a escrita em uma nova planilha com os dados obtidos na Agenda
          (Schedule), ou seja, após o algoritmo de coloração e otimização acontecerem.
          A Classe também é responsável por criar os detalhes mais informativos na
          planilha, definindo cores, bordas, entre outros.
'''

# Classe que lê da Planilha
class Reader:

    # Construtor da classe inicializa o leitor do arquivo
    def __init__(self, arquivo, aba):
        workBook = openpyxl.load_workbook(arquivo)
        self.workSheet = workBook[aba]
        self.maxRow =  self.workSheet.max_row

    # Célula: Captura linha e coluna da planilha
    def getCell(self, linha, coluna):
        return self.workSheet.cell(row = linha, column = coluna).value

    # Percorre a planilha a partir da segunda linha e segunda coluna, devido as
    # informações coletadas no formulário. Faz o tratamento para ' - ' presentes
    # em algumas colunas, criando assim um dicionário
    def buscarPessoas(self):

        pessoas = []

        integer = 2
        for linha in range(2, self.maxRow + 1): # +1 não inclusivo

            # Pega as informações de cada coluna da planilha
            idade = self.getCell(linha, 2)
            sexo = self.getCell(linha, 3)
            cidade = self.getCell(linha, 4)
            quadroDeRisco = self.getCell(linha, 5)
            tempoMedio = self.getCell(linha, 6)

            horarioDePreferencia = str(self.getCell(linha, 7)).split(' - ')
            horarioDePreferencia = toInteger(horarioDePreferencia)

            diasDaSemana = str(self.getCell(linha, 8)).split(' - ')
            diasDaSemana = toInteger(diasDaSemana)

            sintomas = self.getCell(linha, 9)

            # Cria o objeto Pessoa
            pessoa = Pessoa(linha - 1, idade, sexo, cidade, quadroDeRisco, tempoMedio, horarioDePreferencia, diasDaSemana, sintomas)
            # Adiciona a pessoa criada na lista de pessoas e retorna a lista
            pessoas.append(pessoa)

        return pessoas

# Cria o dicionário
def toInteger(lista):
    return list(map(int, lista))

# Classe que Escreve em Planilha
class Writer:
    # Construtor da classe inicializa o escritor do arquivo
    def __init__(self, nomePlanilhaResultado):
        self.nomePlanilhaResultado = nomePlanilhaResultado

    # Gera Planilha alocando os tipos de pessoa com os horários,
    # de acordo com o Algoritmo de Coloração
    def CriaPlanilha(self, schedule, nomeDaAba = 'Todos'):

        # Inicializa o arquivo como nenhum
        arquivo = None

        # Verifica se o arquivo já existe no disco
        arquivoExistente = os.path.isfile(self.nomePlanilhaResultado)

        if arquivoExistente:
            # Carrega o arquivo do disco
            arquivo = openpyxl.load_workbook(filename = self.nomePlanilhaResultado)
            # Apaga a planilha quando o arquivo já existe
            if nomeDaAba in arquivo.sheetnames:
                del arquivo[nomeDaAba]
        else:
            # Cria um novo arquivo
            arquivo = openpyxl.Workbook()
            # Apaga a planilha já configurada ao criar o arquivo
            del arquivo["Sheet"]

        # Define as cores e formato das Bordas
        thin = Side(border_style="thin", color="000000")
        bordas = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Define as cores da célula dos dois grupos: GR e GNR
        patternRisco = PatternFill("solid", fgColor="FFFF99")
        patternNaoRisco = PatternFill("solid", fgColor="99FF99")

        # Define cores das informações adicionais
        fundoPreto = PatternFill("solid", fgColor="000000")
        corDoKellyson = PatternFill("solid", fgColor="777777")

        # Verifica se a aba ainda não existe
        if not (str(nomeDaAba) in arquivo.sheetnames):

            # Cria a nova aba de acordo com o valor por parâmetro
            arquivo.create_sheet(str(nomeDaAba))
            planilha = arquivo[str(nomeDaAba)]

            # Adiciona os Dias da Semana na Planilha
            for i in range(2, 8):
                planilha.cell(row = 1, column = i).value = Enums.DIAS[i - 2]

                # Formata a célula de dias da semana alinhando, tornando negrito,
                # preenchendo o fundo da célula e criando bordas na célula
                planilha.cell(row = 1, column = i).alignment = Alignment(horizontal = 'center', vertical = 'center')
                planilha.cell(row = 1, column = i).fill = PatternFill("solid", fgColor="3366FF")
                planilha.cell(row = 1, column = i).border = bordas
                planilha.cell(row = 1, column = i).font = Font(name='Calibri', bold = True, color="FFFFFF")

            # Adiciona os horários de funcionamento do mercado, de acordo com a pesquisa
            for i in range(2, 13):
                planilha.cell(row = i, column = 1).value = Enums.HORARIOS[i - 2]

                # Formata a célula de horário de funcionamento, tornando negrito,
                # preenchendo o fundo da célula e criando bordas na célula
                planilha.cell(row = i, column = 1).alignment = Alignment(horizontal = 'right', vertical = 'center')
                planilha.cell(row = i, column = 1).fill = PatternFill("solid", fgColor="3366FF")
                planilha.cell(row = i, column = 1).border = bordas
                planilha.cell(row = i, column = 1).font = Font(name='Calibri', bold = True, color="FFFFFF")

            for i in range(1, 13):
                for j in range(1, 7):
                    # Redimensiona as linhas e colunas
                    planilha.row_dimensions[i].height = 40
                    planilha.column_dimensions[Enums.DIA_COLUNA[j - 1]].width = 60


        # Para cada vértice colorido, realiza os processos
        for i, dia in enumerate(schedule.dias):
            for j, horario in enumerate(dia):

                value = ''

                for idx, vertice in enumerate(horario):
                    # Caso exista mais pessoas naquele horário, o símbolo |
                    # fará a separação das pessoas
                    value += self.formatarResposta(vertice.pessoa)
                    if idx < len(horario) - 1:
                        value += ' | '

                # Coloca como valor da célula a lista das pessoas agendadas naquele horário
                # e adiciona borda naquela célula
                planilha.cell(row = j + 2, column = i + 2).value = value
                planilha.cell(row = j + 2, column = i + 2).border = bordas

                # Condição para colorir a célula com
                #   + Amarelo: Pessoas do Grupo de Risco
                #   + Verde: Pessoas não pertencentes ao Grupo de Risco
                if len(horario) > 0:
                    planilha.cell(row = j + 2, column = i + 2).fill = patternRisco if horario[0].pessoa.getQuadroDeRisco() else patternNaoRisco

                # Formata a célula, alinhando-a
                planilha.cell(row = j + 2, column = i + 2).alignment = Alignment(horizontal = 'center', vertical = 'center')

        # Cria uma nova célula abaixo da tabela de horários com a quantidade de
        # cores utilizadas
        planilha.cell(row = 13, column = 7).value = 'Horários ocupados: ' + str(schedule.horariosUtilizados) + '/66'
        planilha.cell(row = 13, column = 7).fill = corDoKellyson
        planilha.cell(row = 13, column = 7).border = bordas
        planilha.cell(row = 13, column = 7).font = Font(name='Calibri', bold = True, color="FFFFFF")

        # Cria nova(s) célula(s) mostrando os horários que foram impossíveis
        # de serem realocados
        if len(schedule.naoAlocados) > 0:
            planilha.cell(row = 14, column = 2).value = 'Os itens abaixo não foram alocados:'
            planilha.cell(row = 14, column = 2).font = Font(name='Calibri', bold = True, color='FFFFFF')
            planilha.cell(row = 14, column = 2).fill = fundoPreto

            for idx, item in enumerate(schedule.naoAlocados):
                planilha.cell(row = 15 + idx, colum = 2).value = self.formatarResposta(item.pessoa)
                planilha.cell(row = 15 + idx, column = 2).border = bordas
                planilha.cell(row = 14, column = 2).font = Font(name='Calibri', bold = True, color='FFFFFF')

        # Cria uma Lista de Planilhas(abas) do arquivo
        ListaDePlanilhas = arquivo.sheetnames

        # Salva o arquivo e o fecha
        arquivo.save(self.nomePlanilhaResultado)

        # Se o arquivo não for fechado, todo esse método não é executado
        arquivo.close()

    # Método que retorna o conteúdo da célula na planilha
    def formatarResposta(self, pessoa):

        string = ''
        string += str(pessoa.index) + " : "
        string += str(pessoa.getIdade()) + " anos, "
        string += pessoa.getCidade()

        return string
