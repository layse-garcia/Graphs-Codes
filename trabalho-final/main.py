# coding: utf-8
#-----------------------------------------------------------------------------------------------------------------------------------#
'''------------------------------------------------------------Imports------------------------------------------------------------'''
# Instalar: pip install openpyxl
import openpyxl

from Pessoa import Pessoa
from Vertice import Vertice

#-----------------------------------------------------------------------------------------------------------------------------------#

def testePlanilha(nomeArquivo, nomeDaAba):
    # Carrego o meu arquivo
    workBook = openpyxl.load_workbook(nomeArquivo)

    workSheet = workBook[nomeDaAba]
    # Planilha começa sempre por 1
    for linha in range(2, workSheet.max_row + 1): # +1 não inclusivo
        idade = workSheet.cell(row = linha, column = 2).value
        sexo = workSheet.cell(row = linha, column = 3).value
        cidade = workSheet.cell(row = linha, column = 4).value
        quadroDeRisco = workSheet.cell(row = linha, column = 5).value
        tempoMedio = workSheet.cell(row = linha, column = 6).value
        horarioDePreferencia = str(workSheet.cell(row = linha, column = 7).value).split(' - ')
        diasDaSemana = str(workSheet.cell(row = linha, column = 8).value).split(' - ')
        sintomas = workSheet.cell(row = linha, column = 9).value

        vertice = Vertice(idade, sexo, cidade, quadroDeRisco, tempoMedio, horarioDePreferencia, diasDaSemana, sintomas)
        print(idade, sexo, cidade, quadroDeRisco, tempoMedio, horarioDePreferencia, diasDaSemana, sintomas)

class Vertice:
  # Construtor do Vertice
  def __init__(self, idade, sexo, cidade, quadroDeRisco, tempoMedioCompras, ListaHorarioCompras, ListaDiasSemana, sintoma):
    self.idade = idade
    self.sexo = sexo
    self.cidade = cidade
    self.quadroDeRisco = quadroDeRisco
    self.tempoMedioCompras = tempoMedioCompras
    self.horarioCompras = ListaHorarioCompras
    self.diaSemana = ListaDiasSemana
    self.sintoma = sintoma
    self.cor = -1

    #FAZ O DESTRUTOR DESSA POHAAA

  def getIdade(self):
    return self.idade

  def getSexo(self):
    return self.sexo

  def getCidade(self):
    return self.cidade

  def getQuadroDeRisco(self):
    return self.quadroDeRisco

  def getTempoMedioCompras(self):
    return self.tempoMedioCompras

  def getHorarioCompras(self):
    return self.horarioCompras

  def getDiaSemana(self):
    return self.diaSemana

  def getSintoma(self):
    return self.sintoma

def main():
    testePlanilha('resultadopesquisa.xlsx', 'respostas')

if __name__ == "__main__":
    main()
